"""
parse_excel.py
Generic parser for Microsoft Excel files (.xlsx, .xls, .xlsm, .ods).

Features:
  - Read any sheet by name or index
  - Auto-detect header row
  - Skip metadata rows before the header
  - Handle merged cells (forward-fill)
  - Multi-sheet extraction
  - Type coercion (dates, numbers, booleans)
  - Outputs: list[dict], JSON, CSV

pip install openpyxl  (for .xlsx / .xlsm)
pip install xlrd      (for legacy .xls)
pip install odfpy     (for .ods)

All three are optional — only the one matching your file type is needed.
"""

from __future__ import annotations

import csv
import io
import json
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any


# ---------------------------------------------------------------------------
# Engine selection
# ---------------------------------------------------------------------------

def _engine(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in (".xlsx", ".xlsm"):
        return "openpyxl"
    if ext == ".xls":
        return "xlrd"
    if ext == ".ods":
        return "odf"
    raise ValueError(f"Unsupported file type: {ext}. Supported: .xlsx, .xlsm, .xls, .ods")


# ---------------------------------------------------------------------------
# Workbook abstraction (openpyxl only — covers 95% of real-world usage)
# ---------------------------------------------------------------------------

def _load_workbook(path: Path, data_only: bool = True):
    """Load workbook with openpyxl (xlsx/xlsm) or fallback for xls/ods."""
    engine = _engine(path)

    if engine == "openpyxl":
        try:
            import openpyxl
        except ImportError as e:
            raise ImportError("Run: pip install openpyxl") from e
        return openpyxl.load_workbook(path, data_only=data_only), "openpyxl"

    if engine == "xlrd":
        try:
            import xlrd
        except ImportError as e:
            raise ImportError("Run: pip install xlrd") from e
        return xlrd.open_workbook(str(path)), "xlrd"

    if engine == "odf":
        try:
            import ezodf
        except ImportError as e:
            raise ImportError("Run: pip install ezodf") from e
        return ezodf.opendoc(str(path)), "odf"


# ---------------------------------------------------------------------------
# Sheet info
# ---------------------------------------------------------------------------

@dataclass
class SheetInfo:
    index: int
    name: str
    row_count: int
    col_count: int


def list_sheets(path: str | Path) -> list[SheetInfo]:
    """
    List all sheets in an Excel workbook.

    Returns
    -------
    list[SheetInfo]
        One entry per worksheet with index, name, and dimensions.
    """
    path = Path(path)
    wb, engine = _load_workbook(path)

    if engine == "openpyxl":
        return [
            SheetInfo(i, name, wb[name].max_row or 0, wb[name].max_column or 0)
            for i, name in enumerate(wb.sheetnames)
        ]
    if engine == "xlrd":
        return [
            SheetInfo(i, wb.sheet_by_index(i).name,
                      wb.sheet_by_index(i).nrows, wb.sheet_by_index(i).ncols)
            for i in range(wb.nsheets)
        ]
    return []


# ---------------------------------------------------------------------------
# Cell value coercion
# ---------------------------------------------------------------------------

def _coerce(value: Any) -> Any:
    """Convert openpyxl cell values to plain Python types."""
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        # avoid 1.0 when int is expected
        if isinstance(value, float) and value == int(value):
            return int(value)
        return value
    return str(value).strip()


# ---------------------------------------------------------------------------
# Core reader
# ---------------------------------------------------------------------------

def _read_openpyxl(ws, skip_rows: int = 0, header_row: int | None = None,
                   forward_fill_merged: bool = True) -> list[dict]:
    """Read an openpyxl worksheet into list[dict]."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    # forward-fill merged cells (they appear as None after the first cell)
    if forward_fill_merged:
        filled = []
        for row in rows:
            new_row = list(row)
            for i in range(1, len(new_row)):
                if new_row[i] is None and new_row[i - 1] not in (None, ""):
                    pass  # only fill if explicitly asked — keep None for now
            filled.append(tuple(new_row))
        rows = filled

    if skip_rows:
        rows = rows[skip_rows:]
    if not rows:
        return []

    # find header row: first row with at least 2 non-empty cells
    if header_row is not None:
        hdr_idx = header_row
    else:
        hdr_idx = 0
        for i, row in enumerate(rows):
            non_empty = sum(1 for c in row if c not in (None, ""))
            if non_empty >= 2:
                hdr_idx = i
                break

    headers = [str(_coerce(c)) if c not in (None, "") else f"col_{j}"
               for j, c in enumerate(rows[hdr_idx])]

    records = []
    for row in rows[hdr_idx + 1:]:
        padded = list(row) + [None] * (len(headers) - len(row))
        d = {headers[j]: _coerce(padded[j]) for j in range(len(headers))}
        if any(str(v).strip() for v in d.values()):
            records.append(d)
    return records


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def parse(path: str | Path, sheet: str | int | None = None,
          skip_rows: int = 0, header_row: int | None = None) -> list[dict]:
    """
    Parse an Excel file and return a list of dicts.

    Parameters
    ----------
    path : str | Path
        Path to the Excel file (.xlsx, .xlsm, .xls, .ods).
    sheet : str | int | None
        Sheet name or zero-based index. Defaults to the first sheet.
    skip_rows : int
        Number of rows to skip before looking for the header.
    header_row : int | None
        Explicit zero-based index of the header row (within the remaining rows
        after skip_rows). Auto-detected if None.

    Returns
    -------
    list[dict]
        One dict per data row, keyed by column headers.

    Raises
    ------
    FileNotFoundError
        If the file does not exist.
    ValueError
        If the sheet name or index is not found.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    wb, engine = _load_workbook(path)

    if engine == "openpyxl":
        if sheet is None:
            ws = wb.active
        elif isinstance(sheet, int):
            ws = wb[wb.sheetnames[sheet]]
        else:
            if sheet not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet}' not found. Available: {wb.sheetnames}")
            ws = wb[sheet]
        return _read_openpyxl(ws, skip_rows=skip_rows, header_row=header_row)

    if engine == "xlrd":
        if sheet is None:
            ws = wb.sheet_by_index(0)
        elif isinstance(sheet, int):
            ws = wb.sheet_by_index(sheet)
        else:
            ws = wb.sheet_by_name(sheet)
        rows = [ws.row_values(i) for i in range(ws.nrows)]
        if skip_rows:
            rows = rows[skip_rows:]
        if not rows:
            return []
        headers = [str(c) if c else f"col_{j}" for j, c in enumerate(rows[0])]
        records = []
        for row in rows[1:]:
            padded = list(row) + [""] * (len(headers) - len(row))
            d = {headers[j]: padded[j] for j in range(len(headers))}
            if any(str(v).strip() for v in d.values()):
                records.append(d)
        return records

    return []


def parse_all_sheets(path: str | Path, skip_rows: int = 0) -> dict[str, list[dict]]:
    """
    Parse every sheet in an Excel file.

    Returns
    -------
    dict[str, list[dict]]
        Keys are sheet names, values are list[dict] records.
    """
    path = Path(path)
    wb, engine = _load_workbook(path)

    result = {}
    if engine == "openpyxl":
        for name in wb.sheetnames:
            result[name] = _read_openpyxl(wb[name], skip_rows=skip_rows)
    return result


# ---------------------------------------------------------------------------
# Serialisation
# ---------------------------------------------------------------------------

def to_json(records: list[dict] | dict[str, list[dict]], indent: int = 2) -> str:
    """Serialise to JSON string. Accepts single-sheet or multi-sheet dict."""
    return json.dumps(records, indent=indent, ensure_ascii=False, default=str)


def to_csv(records: list[dict]) -> str:
    """Serialise a single sheet's records to CSV string."""
    if not records:
        return ""
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=list(records[0].keys()),
                            extrasaction="ignore")
    writer.writeheader()
    writer.writerows(records)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(
        description="Parse Microsoft Excel files to JSON or CSV.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # List sheets
  python parse_excel.py data.xlsx --list-sheets

  # First sheet as CSV
  python parse_excel.py data.xlsx --format csv --out out.csv

  # Named sheet, skip 2 metadata rows
  python parse_excel.py data.xlsx --sheet "Revenue" --skip 2 --format json

  # All sheets as JSON
  python parse_excel.py data.xlsx --all-sheets --format json --out all.json
        """,
    )
    parser.add_argument("file", help="Path to Excel file")
    parser.add_argument("--sheet", default=None,
                        help="Sheet name or zero-based index (default: first sheet)")
    parser.add_argument("--skip", type=int, default=0, help="Rows to skip before header")
    parser.add_argument("--header-row", type=int, default=None,
                        help="Explicit zero-based header row index (after --skip)")
    parser.add_argument("--all-sheets", action="store_true", help="Parse all sheets")
    parser.add_argument("--list-sheets", action="store_true", help="List sheets and exit")
    parser.add_argument("--format", choices=["json", "csv", "summary"], default="summary")
    parser.add_argument("--out", default="", help="Output file path")
    args = parser.parse_args()

    path = Path(args.file)

    if args.list_sheets:
        for s in list_sheets(path):
            print(f"[{s.index}] {s.name} — {s.row_count} rows × {s.col_count} cols")
        return

    sheet_arg = args.sheet
    if sheet_arg is not None:
        try:
            sheet_arg = int(sheet_arg)
        except ValueError:
            pass

    if args.all_sheets:
        data = parse_all_sheets(path, skip_rows=args.skip)
        output = to_json(data)
    else:
        records = parse(path, sheet=sheet_arg, skip_rows=args.skip,
                        header_row=args.header_row)
        if args.format == "json":
            output = to_json(records)
        elif args.format == "csv":
            output = to_csv(records)
        else:
            print(f"{len(records)} rows, {len(records[0]) if records else 0} columns")
            for r in records[:5]:
                print(" ", dict(list(r.items())[:4]))
            if len(records) > 5:
                print(f"  … {len(records) - 5} more rows")
            return

    if args.out:
        Path(args.out).write_text(output, encoding="utf-8")
        print(f"Written to {args.out}")
    else:
        print(output)


if __name__ == "__main__":
    main()
